from django.contrib.auth.mixins import LoginRequiredMixin
from django.http.response import HttpResponse
from django_filters.views import FilterView
from django.contrib import messages
from django.urls.base import reverse_lazy
from django.views.generic import *
from openpyxl import Workbook

from .time_manager import *
from .models import *
from .filters import *
from MainApp.models import *
from .forms import *
from .utils import *

class Operations(LoginRequiredMixin, FilterView):
    template_name = "MoneyControlApp/operations.html"
    model = Operation
    login_url = reverse_lazy("login")
    filterset_class = OperationsFilter

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Операции"
        context["bal"] = Operation.get_sum(self.object_list)
        return context

    def get_queryset(self):
        # Показывает пользователю только его операции
        return Operation.objects.filter(user=self.request.user).order_by('-date')

class AddOperation(LoginRequiredMixin, CreateView):
    form_class = OperationForm
    template_name = 'MainApp/form.html'
    login_url = reverse_lazy('login')
    success_url = reverse_lazy('operations')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Добавить операцию"
        context["h3_text"] = "Добавить операцию"
        context["button_text"] = "Добавить"
        return context
    
    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.user = self.request.user
        self.object.save()
        messages.add_message(self.request, messages.SUCCESS, 'Вы успешно добавили операцию.')
        return super().form_valid(form)

class EditOperation(LoginRequiredMixin, UpdateView):
    model = Operation
    form_class = OperationForm
    pk_url_kwarg = "operation_id"
    template_name = 'MainApp/form.html'
    login_url = reverse_lazy('login')
    success_url = reverse_lazy('operations')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Изменить операцию"
        context["h3_text"] = "Изменить операцию"
        context["button_text"] = "Сохранить"
        context["categories"] = [(cat.id, cat.name, cat.type) for cat in Category.objects.all()]
        return context

    def form_valid(self, form):
        messages.add_message(self.request, messages.SUCCESS, 'Вы успешно изменили операцию.')
        return super().form_valid(form)
    
    def get_queryset(self):
        # Пользователь может изменять только свои расходы
        return Operation.objects.filter(user=self.request.user)

class DeleteOperation(LoginRequiredMixin, DeleteView):
    model = Operation
    template_name = 'MainApp/form.html'
    pk_url_kwarg = "operation_id"
    login_url = reverse_lazy('login')
    success_url = reverse_lazy('operations')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Удалить операцию"
        context["h3_text"] = "Удалить операцию"
        context["button_text"] = "Да"
        context["text"] = "Вы точно хотите операцию?"
        return context

    def post(self, request, *args, **kwargs):
        messages.add_message(request, messages.SUCCESS, 'Вы успешно удалили операцию.')
        return super().post(request, *args, **kwargs)

    def get_queryset(self):
        # Пользователь может удалять только свои операции
        return Operation.objects.filter(user=self.request.user)

class Categories(LoginRequiredMixin, FilterView):
    template_name = "MoneyControlApp/categories.html"
    model = Operation
    filterset_class = BaseOperationFilter
    login_url = reverse_lazy("login")

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Категории"
        operations = self.object_list
        
        # Прогноз расходов до конца месяца
        forecast = ExpenseForecast()
        value, month_date = forecast.createForecast(self.qs)
        context["forecast"] = value
        context["month_data"] = month_date

        # Расчет остатка на будущее с использованием ИПЦ
        cpi = CPI()
        bal = Operation.get_sum(self.qs)

        cpi_3_month, cpi_6_month, cpi_year = cpi.getNextTreeMonth(bal), cpi.getNextSixMonth(bal), cpi.getNextYear(bal)
        print(cpi_3_month, cpi_year, cpi_6_month)

        cpi_data = [("3 месяца", round(cpi_3_month * bal / 100), round(sum(cpi.cpi_data[:3]) / 3 - 100, 3)), 
                    ("6 месяцев", round(cpi_6_month * bal / 100), round(sum(cpi.cpi_data[:6]) / 6 - 100, 3)),
                    ("год", round(cpi_year * bal / 100), round(sum(cpi.cpi_data) / 12 - 100, 3))]

        context["cpi_data"] = cpi_data
        context["bal"] = bal

        # Данные о категориях
        data = {}
        for operation in operations:
            if operation.category.name not in data:
                data[operation.category.name] = (operation.amount)
            else:
                data[operation.category.name] += operation.amount
    
        total_amount = sum(data.values())

        cat_data = []
        for elem in data:
            percent = round(data[elem] * 100 / total_amount, 2)
            cat_data.append([elem, data[elem], percent])

        # Сортируем по общей сумме операций этой категории
        cat_data = sorted(cat_data, key=lambda cat: -cat[1])
        context["cat_data"] = cat_data

        return context

    def get_queryset(self):
        qs = super().get_queryset().filter(user=self.request.user)
        self.qs = qs
        return qs

class ExportData(LoginRequiredMixin, FormView):
    model = Operation
    form_class = ExportDataForm
    template_name = 'MainApp/form.html'
    login_url = reverse_lazy('login')
    success_url = reverse_lazy('home')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Экспорт данных"
        context["h3_text"] = "Перевод данных в .xlsx"
        context["button_text"] = "Скачать"
        return context

    def form_valid(self, form):
        # Получаем данные из формы и сортируем операции
        date_begin = form.cleaned_data['date_begin']
        date_end = form.cleaned_data['date_end']
        operations = Operation.objects.filter(user=self.request.user, date__gte=date_begin, date__lte=date_end)
        
        TABLE_HEAD = ["Тип", "Сумма", "Дата", "Категория"]

        # Создаем таблицу
        workbook = Workbook()
        sheet = workbook.active

        # Настраиваем таблицу
        sheet.column_dimensions['C'].width = 13
        sheet.column_dimensions['D'].width = 13

        # Создаем шапку таблицы
        for col, val in enumerate(TABLE_HEAD):
            sheet.cell(row=1, column=col + 1, value=val)
        
        # Вставляем данные в таблицу
        for row, operation in enumerate(operations):
            values = [operation.category.type, operation.amount, operation.date, operation.category.name]
            for col, val in enumerate(values):
                sheet.cell(row=row + 2, column=col + 1, value=val)
        
        # Подготовка файла для передачи пользователю
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=operations.xlsx'
        workbook.save(response)

        return response
